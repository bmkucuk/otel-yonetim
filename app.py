#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Otel Leo & Cunda Villa — Web Yönetim (SQLite sürümü)"""
import os
from datetime import date
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
import hashlib
from functools import wraps
import database as db
from muhasebe_routes import muh
import muhasebe_db as mdb

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'otelleo2026x!9k')

# ── Kullanıcılar ──────────────────────────────────────────────────────────────
USERS = {
    'bmkucuk@gmail.com': {'hash': '23b5c5e0915483302bbd48e555a85f5999f52dbed8c7c7a5809811bba234e0a1', 'role': 'admin'},
    'villacunda':        {'hash': '29624e2e4c4ccee26ed8f3e0ca1012ea57a8f2191be6149f632250f7036119cc', 'role': 'personel'},
}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect('/login')
        if session.get('role') != 'admin':
            return redirect('/')
        return f(*args, **kwargs)
    return decorated

# Muhasebe blueprint
app.register_blueprint(muh)

# DB başlat
db.init_db()
mdb.init_db()

# ── Auth ─────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def login():
    if session.get('user'):
        return redirect('/')
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        h = hashlib.sha256(password.encode()).hexdigest()
        user = USERS.get(username)
        if user and user['hash'] == h:
            session['user'] = username
            session['role'] = user['role']
            return redirect('/')
        return render_template('login.html', error='Kullanıcı adı veya şifre yanlış')
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


# ── Sayfalar ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/rezervasyonlar')
@login_required
def rezervasyonlar():
    return render_template('rezervasyonlar.html')

@app.route('/oda-durumu')
@login_required
def oda_durumu():
    return render_template('oda_durumu.html')

@app.route('/cari')
@login_required
def cari():
    return render_template('cari.html')

@app.route('/adisyon')
@login_required
def adisyon():
    return render_template('adisyon.html')

@app.route('/gunluk-liste')
@login_required
def gunluk_liste():
    return render_template('gunluk_liste.html')

@app.route('/import')
@admin_required
def import_page():
    return render_template('import.html')

@app.route('/yedekleme-talimati')
@admin_required
def yedekleme_talimati():
    return render_template('yedekleme_talimati.html')

@app.route('/api/gunluk-liste')
@login_required
def api_gunluk_liste():
    tarih = request.args.get('tarih', date.today().isoformat())
    conn = db.get_conn()
    # Giriş listesi
    girisler = conn.execute(
        "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no "
        "FROM rezervasyonlar WHERE giris=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY otel, oda_no", (tarih,)
    ).fetchall()
    # Çıkış listesi
    cikislar = conn.execute(
        "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no "
        "FROM rezervasyonlar WHERE cikis=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY otel, oda_no", (tarih,)
    ).fetchall()
    # Kahvaltı listesi: konaklıyorlar (giris < tarih <= cikis) VEYA çıkış günü (cikis = tarih)
    # Giriş günü kahvaltı YOK, çıkış günü VAR
    kahvalti = conn.execute(
        "SELECT oda_no, otel, musteri, yetiskin, cocuk, giris, cikis, foy_no "
        "FROM rezervasyonlar "
        "WHERE giris < ? AND cikis >= ? AND (durum IS NULL OR durum != 'Kapora Yandı') "
        "ORDER BY otel, oda_no", (tarih, tarih)
    ).fetchall()
    conn.close()
    def row2dict(r):
        return {
            'oda_no': r['oda_no'], 'otel': r['otel'], 'musteri': r['musteri'],
            'yetiskin': r['yetiskin'], 'cocuk': r['cocuk'],
            'giris': r['giris'], 'cikis': r['cikis'], 'foy_no': r['foy_no']
        }
    return jsonify({
        'tarih': tarih,
        'girisler': [row2dict(r) for r in girisler],
        'cikislar': [row2dict(r) for r in cikislar],
        'kahvalti': [row2dict(r) for r in kahvalti]
    })


# ── API — Okuma ───────────────────────────────────────────────────────────────

@app.route('/api/dashboard')
def api_dashboard():
    today = date.today().isoformat()
    girisler, cikislar, aktifler, kahvalti = db.get_dashboard(today)
    leo_aktif = sum(1 for r in aktifler if r['otel'] == 'LEO')
    cv_aktif  = sum(1 for r in aktifler if r['otel'] == 'CV')
    toplam_alacak = sum((r.get('rez_bakiye') or 0) + (r.get('adis_bakiye') or 0)
                        for r in db.get_rezervasyonlar()
                        if r.get('durum') != 'Kapora Yandı')
    return jsonify({
        'today': date.today().strftime('%d %B %Y, %A'),
        'stats': {
            'bugun_giris':   len(girisler),
            'bugun_cikis':   len(cikislar),
            'leo_aktif':     leo_aktif,
            'cv_aktif':      cv_aktif,
            'kahvalti_kisi': kahvalti,
            'toplam_alacak': toplam_alacak,
        },
        'girisler': girisler,
        'cikislar': cikislar,
        'aktifler': aktifler,
    })

@app.route('/api/rezervasyonlar')
def api_rezervasyonlar():
    q     = request.args.get('q', '')
    otel  = request.args.get('otel', 'Tümü')
    return jsonify(db.get_rezervasyonlar(q=q, otel=otel))

@app.route('/api/oda-durumu')
def api_oda_durumu():
    from datetime import timedelta
    start_str = request.args.get('start', date.today().isoformat())
    try:
        start = date.fromisoformat(start_str)
    except:
        start = date.today()
    days  = 14
    dates = [start + timedelta(i) for i in range(days)]
    today = date.today().isoformat()

    rezervasyonlar = [r for r in db.get_rezervasyonlar() if r.get("durum") != "Kapora Yandı"]
    cv_odalar  = list(range(1, 11))
    leo_odalar = list(range(11, 30))
    all_rooms  = [('CV', o) for o in cv_odalar] + [('LEO', o) for o in leo_odalar]

    grid = []
    for otel_label, oda_no in all_rooms:
        cells = []
        for d in dates:
            ds = d.isoformat()
            rez = next((r for r in rezervasyonlar
                        if r['oda_no'] == oda_no
                        and r['giris'] and r['cikis']
                        and r['giris'] <= ds < r['cikis']), None)
            if rez:
                parts = rez['musteri'].split()
                initials = ' '.join(p[0] for p in parts[:2]) if parts else '?'
                cells.append({
                    'dolu': True, 'initials': initials,
                    'musteri': rez['musteri'],
                    'giris': rez['giris'], 'cikis': rez['cikis'],
                    'foy_no': rez['foy_no'],
                    'is_giris': ds == rez['giris'],
                    'otel': rez['otel'],
                })
            else:
                cells.append({'dolu': False})
        grid.append({'otel': otel_label, 'oda_no': oda_no, 'cells': cells})

    return jsonify({'dates': [d.isoformat() for d in dates], 'today': today, 'grid': grid})

@app.route('/api/cari')
def api_cari():
    rezervasyonlar = db.get_rezervasyonlar()
    def _f(v): return float(v or 0)
    aktif_rez = [r for r in rezervasyonlar if r.get('durum') != 'Kapora Yandı']
    ozet = {
        'toplam_rez':      sum(_f(r.get('toplam_fiyat')) for r in aktif_rez),
        'toplam_kapora':   sum(_f(r.get('kapora')) for r in aktif_rez),
        'toplam_tahsilat': sum(_f(r.get('rez_tahsilat')) for r in aktif_rez),
        'rez_bakiye':      sum(_f(r.get('rez_bakiye')) for r in aktif_rez),
        'adis_toplam':     sum(_f(r.get('adisyon')) for r in aktif_rez),
        'adis_tahsilat':   sum(_f(r.get('adis_tahsilat')) for r in aktif_rez),
        'adis_bakiye':     sum(_f(r.get('adis_bakiye')) for r in aktif_rez),
    }
    return jsonify({'ozet': ozet, 'rezervasyonlar': rezervasyonlar})

@app.route('/api/adisyonlar')
def api_adisyonlar():
    foy_f = request.args.get('foy')
    adisyonlar = db.get_adisyonlar(foy_no=int(foy_f) if foy_f else None)
    rezervasyonlar = db.get_rezervasyonlar()
    musteri_map = {r['foy_no']: r['musteri'] for r in rezervasyonlar}
    for a in adisyonlar:
        a['musteri'] = musteri_map.get(a['foy_no'], '')
    aktif_rez = [r for r in rezervasyonlar if r.get('durum') != 'Kapora Yandı']
    foy_listesi = [{'foy_no': r['foy_no'], 'musteri': r['musteri'], 'oda_no': r['oda_no']}
                   for r in sorted(aktif_rez, key=lambda x: x['foy_no'])]
    return jsonify({'adisyonlar': adisyonlar, 'foy_listesi': foy_listesi})

@app.route('/api/next-foy')
def api_next_foy():
    return jsonify({'foy_no': db.get_next_foy_no()})

@app.route('/api/next-adisyon')
def api_next_adisyon():
    return jsonify({'adisyon_no': db.get_next_adisyon_no()})


# ── API — Yazma ───────────────────────────────────────────────────────────────

@app.route('/api/rezervasyon/yeni', methods=['POST'])
def api_rez_yeni():
    try:
        data = request.get_json()
        db.save_rezervasyon(data)
        # Yevmiye: konaklama geliri
        yevmiye_rez_kaydet(
            int(data.get('foy_no') or 0), float(data.get('toplam_fiyat') or 0),
            data.get('otel', 'LEO'), data.get('giris'),
            data.get('musteri', ''),
            kapora=float(data.get('kapora') or 0),
            kapora_tarihi=data.get('kapora_tarihi')
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/sil', methods=['POST'])
def api_rez_sil():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        # Yevmiyeden ilgili kayıtları sil
        conn = mdb.get_conn()
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ?", (f'%Föy#{foy_no}%',))
        conn.commit(); conn.close()
        # Rezervasyonu sil
        db.delete_rezervasyon(foy_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/checkin', methods=['POST'])
@login_required
def api_checkin():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        if not r:
            return jsonify({'ok': False, 'error': 'Rezervasyon bulunamadı'}), 404
        if r.get('checkin'):
            return jsonify({'ok': False, 'error': 'Check-in zaten yapılmış'}), 400

        otel = r.get('otel', 'LEO')
        musteri = r.get('musteri', '')
        toplam = float(r.get('toplam_fiyat') or 0)
        kapora = float(r.get('kapora') or 0)
        tarih = date.today().isoformat()
        gelir_hesap = '600' if otel == 'LEO' else '601'
        aciklama = f'Föy#{foy_no} {musteri} check-in'

        # Check-in kaydını güncelle
        conn_db = db.get_conn()
        conn_db.execute("UPDATE rezervasyonlar SET checkin=1, durum='Konaklıyor' WHERE foy_no=?", (foy_no,))
        conn_db.commit(); conn_db.close()

        # Yevmiye:
        # A. Gelir doğumu: 120 borç / 601 alacak (toplam konaklama)
        # B. Kapora mahsubu: 340 borç / 120 alacak
        conn = mdb.get_conn()
        if toplam > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Check-in Konaklama', '120', gelir_hesap,
                              toplam, aciklama, otel)
        if kapora > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Kapora Mahsubu', '340', '120',
                              kapora, aciklama, otel)
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/checkin-iptal', methods=['POST'])
@login_required
def api_checkin_iptal():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        if not r or not r.get('checkin'):
            return jsonify({'ok': False, 'error': 'Check-in bulunamadı'}), 400

        otel = r.get('otel', 'LEO')
        musteri = r.get('musteri', '')
        toplam = float(r.get('toplam_fiyat') or 0)
        kapora = float(r.get('kapora') or 0)
        tarih = date.today().isoformat()
        gelir_hesap = '600' if otel == 'LEO' else '601'
        aciklama = f'Föy#{foy_no} {musteri} check-in iptal (storno)'

        # Check-in iptal
        conn_db = db.get_conn()
        conn_db.execute("UPDATE rezervasyonlar SET checkin=0, durum='Aktif' WHERE foy_no=?", (foy_no,))
        conn_db.commit(); conn_db.close()

        # Storno kayıtları (ters çevir)
        conn = mdb.get_conn()
        if toplam > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Check-in Storno', gelir_hesap, '120',
                              toplam, aciklama, otel)
        if kapora > 0:
            mdb._yevmiye_ekle(conn, tarih, 'Kapora Mahsubu Storno', '120', '340',
                              kapora, aciklama, otel)
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/rezervasyon/kapora-yandi', methods=['POST'])
def api_kapora_yandi():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        rezler = db.get_rezervasyonlar()
        r = next((x for x in rezler if x['foy_no'] == foy_no), None)
        if not r:
            return jsonify({'ok': False, 'error': 'Rezervasyon bulunamadı'}), 404
        kapora = float(r.get('kapora') or 0)
        if kapora <= 0:
            return jsonify({'ok': False, 'error': 'Kapora yok'}), 400
        otel = r.get('otel', 'LEO')
        gelir_hesap = '600' if otel == 'LEO' else '601'
        musteri = r.get('musteri', '')
        tarih = date.today().isoformat()
        # Yevmiye: Alınan Avanslar borç / Diğer Olağan Gelir alacak (340/649)
        conn = mdb.get_conn()
        mdb._yevmiye_ekle(conn, tarih, 'Kapora Yanması', '340', '649',
                          kapora, f'Föy#{foy_no} {musteri} kapora yandı - iptal bedeli', otel)
        conn.commit(); conn.close()
        # Rezervasyon durumunu güncelle, bakiyeyi sıfırla (silinmez)
        otel_conn = db.get_conn()
        otel_conn.execute(
            "UPDATE rezervasyonlar SET durum='Kapora Yandı', rez_bakiye=0 WHERE foy_no=?",
            (foy_no,)
        )
        otel_conn.commit(); otel_conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


@app.route('/api/rezervasyon/guncelle', methods=['POST'])
def api_rez_guncelle():
    try:
        data = request.get_json()
        db.update_rezervasyon(data['foy_no'], data)
        # Yevmiye: konaklama geliri güncelle
        yevmiye_rez_kaydet(
            int(data.get('foy_no') or 0), float(data.get('toplam_fiyat') or 0),
            data.get('otel', 'LEO'), data.get('giris'),
            data.get('musteri', ''),
            kapora=float(data.get('kapora') or 0),
            kapora_tarihi=data.get('kapora_tarihi'),
            guncelleme=True
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


def yevmiye_rez_kaydet(foy_no, toplam_fiyat, otel, giris, musteri,
                       kapora=0, kapora_tarihi=None, guncelleme=False):
    """Rezervasyon konaklama geliri ve kaporayı yevmiyeye yazar."""
    try:
        tarih = giris or date.today().isoformat()
        gelir_hesap = '600' if otel == 'LEO' else '601'
        aciklama_kon = f'Föy#{foy_no} {musteri} konaklama'
        aciklama_kap = f'Föy#{foy_no} {musteri} kapora'
        conn = mdb.get_conn()
        # Eski kayıtları sil - foy_no ile eşleştir (misafir adı değişse bile çalışır)
        conn.execute("DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi IN ('Konaklama Geliri','Kapora')",
                     (f'Föy#{foy_no} %',))
        conn.commit()

        kap_tarih = kapora_tarihi or tarih

        # Rezervasyon anında sadece kapora: İş Bankası borç / Alınan Avanslar alacak (340)
        # Konaklama geliri check-in anında yazılacak
        if kapora and kapora > 0:
            mdb._yevmiye_ekle(conn, kap_tarih, 'Kapora', '102-1', '340',
                              kapora, aciklama_kap, otel)

        conn.commit(); conn.close()
    except Exception as e:
        print(f'Yevmiye rez kayıt hatası: {e}')

ODEME_HESAP_KODU = {
    'Nakit': '100', 'nakit': '100',
    'Kredi Kartı': '102-1', 'KK': '102-1', 'kk': '102-1',
    'Havale': '102-1', 'EFT': '102-1',
}



@app.route('/api/tahsilat/gecmis')
def api_tahsilat_gecmis():
    foy_no = request.args.get('foy_no', type=int)
    conn = mdb.get_conn()
    rows = conn.execute("""
        SELECT id, tarih, islem_tipi, tutar, borc_hesap
        FROM yevmiye 
        WHERE (islem_tipi LIKE '%Tahsilat%' OR islem_tipi = 'Kapora' OR islem_tipi = 'Kapora Yanması') 
        AND aciklama LIKE ?
        ORDER BY tarih ASC
    """, (f'%Föy#{foy_no}%',)).fetchall()
    conn.close()
    odeme_map = {
        '100': 'Nakit', '102-1': 'İş Bankası', '102-2': 'Ziraat',
        '102-3': 'Deniz Bank', '120': 'Müşteri Cari', '340': 'Alınan Kaparo'
    }
    result = []
    for r in rows:
        result.append({
            'id': r[0], 'tarih': r[1], 'tur': r[2],
            'tutar': r[3], 'odeme': odeme_map.get(r[4], r[4]),
            'foy_no': foy_no
        })
    return jsonify(result)

@app.route('/api/tahsilat/sil', methods=['POST'])
def api_tahsilat_sil():
    try:
        d = request.get_json()
        yev_id = d['yev_id']
        foy_no = int(d['foy_no'])
        tutar = float(d['tutar'])
        tur = d['tur']
        
        # Yevmiyeyi sil
        conn = mdb.get_conn()
        conn.execute("DELETE FROM yevmiye WHERE id=?", (yev_id,))
        conn.commit()
        conn.close()
        
        # Rezervasyon tablosunu güncelle
        otel_conn = db.get_conn()
        if 'Adisyon' in tur:
            otel_conn.execute("""
                UPDATE rezervasyonlar 
                SET adis_tahsilat = adis_tahsilat - ?,
                    adis_bakiye = adis_bakiye + ?
                WHERE foy_no=?
            """, (tutar, tutar, foy_no))
        else:
            otel_conn.execute("""
                UPDATE rezervasyonlar 
                SET rez_tahsilat = rez_tahsilat - ?,
                    rez_bakiye = rez_bakiye + ?
                WHERE foy_no=?
            """, (tutar, tutar, foy_no))
        otel_conn.commit()
        otel_conn.close()
        
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
@app.route('/api/tahsilat/rez', methods=['POST'])
def api_rez_tah():
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        tutar = float(d['tutar'])
        odeme = d['odeme']
        db.save_rez_tahsilat(foy_no, tutar, odeme)
        # Yevmiye: otomatik kayıt
        hesap_kodu = ODEME_HESAP_KODU.get(odeme)
        if hesap_kodu and tutar > 0:
            rez = db.get_rezervasyonlar()
            r = next((x for x in rez if x['foy_no'] == foy_no), None)
            tarih = date.today().isoformat()
            otel = r.get('otel', 'LEO') if r else 'LEO'
            musteri = r.get('musteri', '') if r else ''
            gelir_hesap = '600' if otel == 'LEO' else '601'
            conn = mdb.get_conn()
            mdb._yevmiye_ekle(conn, tarih, 'Rezervasyon Tahsilat - ' + odeme,
                              hesap_kodu, '120', tutar,
                              f'Föy#{foy_no} {musteri} rez tahsilat', otel)
            # Kapora mahsubu check-in anında yazılıyor, tahsilatta yazılmaz
            conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/tahsilat/adis', methods=['POST'])
def api_adis_tah():
    """Adisyon bazlı kısmi veya tam ödeme."""
    try:
        d = request.get_json()
        foy_no = int(d['foy_no'])
        odeme = d['odeme']
        adisyon_nolar = d.get('adisyon_nolar', [])  # seçili adisyon no listesi
        # Her adisyon için tutar ayrıca gönderilir: {adisyon_no: tutar}
        adis_tutarlar = d.get('adis_tutarlar', {})

        rez = db.get_rezervasyonlar()
        r = next((x for x in rez if x['foy_no'] == foy_no), None)
        otel = r.get('otel', 'LEO') if r else 'LEO'
        musteri = r.get('musteri', '') if r else ''

        toplam_odeme = 0
        for adis_no in adisyon_nolar:
            tutar = float(adis_tutarlar.get(str(adis_no), 0))
            if tutar <= 0:
                continue
            toplam_odeme += tutar
            # Adisyon bazlı ödeme kaydet
            tarih = db.save_adisyon_odeme(adis_no, foy_no, tutar, odeme)

        if toplam_odeme > 0:
            # Rezervasyon tablosundaki adis_tahsilat güncelle
            db.save_adis_tahsilat(foy_no, toplam_odeme, odeme)

            # Yevmiye - her adisyon için 2 kayıt:
            # 1. hesap_kodu borç / 120 alacak (para geldi, müşteri borcu kapandı)
            # 2. 120 borç / 610 alacak (adisyon geliri gerçekleşti)
            hesap_kodu = ODEME_HESAP_KODU.get(odeme)
            if hesap_kodu:
                tarih = date.today().isoformat()
                conn = mdb.get_conn()
                for adis_no in adisyon_nolar:
                    t = float(adis_tutarlar.get(str(adis_no), 0))
                    if t > 0:
                        aciklama = f'Föy#{foy_no} Adis#{adis_no} {musteri} adisyon tahsilat'
                        # Para geldi: Kasa/Banka borç / Müşteri Cari alacak
                        mdb._yevmiye_ekle(conn, tarih, 'Adisyon Tahsilat - ' + odeme,
                                          hesap_kodu, '120', t, aciklama, otel)
                conn.commit(); conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/odemeler', methods=['GET'])
def api_adisyon_odemeler():
    """Bir adisyona ait ödeme geçmişi."""
    adisyon_no = request.args.get('adisyon_no', type=int)
    if not adisyon_no:
        return jsonify([])
    return jsonify(db.get_adisyon_odemeler(adisyon_no))

@app.route('/api/adisyon/ekle', methods=['POST'])
def api_adisyon_ekle():
    try:
        d = request.get_json()
        db.save_adisyon(d)
        # Yevmiye: Müşteri Cari borç / Adisyon Geliri alacak
        tutar = float(d.get('tutar') or 0)
        foy_no = int(d.get('foy_no') or 0)
        adisyon_no = int(d.get('adisyon_no') or 0)
        if tutar > 0 and foy_no:
            rez = db.get_rezervasyonlar()
            r = next((x for x in rez if x['foy_no'] == foy_no), None)
            otel = r.get('otel', 'GENEL') if r else 'GENEL'
            musteri = r.get('musteri', '') if r else ''
            tarih = d.get('tarih') or date.today().isoformat()
            conn = mdb.get_conn()
            mdb._yevmiye_ekle(conn, tarih, 'Adisyon Geliri', '120', '610',
                              tutar, f'Föy#{foy_no} Adis#{adisyon_no} {musteri} adisyon', otel)
            conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/guncelle', methods=['POST'])
def api_adisyon_guncelle():
    try:
        d = request.get_json()
        adisyon_no = int(d['adisyon_no'])
        yeni_tutar = float(d['tutar'])
        odeme = d.get('odeme', 'Oda Hesabına')

        # Eski adisyon bilgisini al
        conn_db = db.get_conn()
        a = conn_db.execute("SELECT * FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
        conn_db.close()

        if a:
            eski_tutar = float(a['tutar'] or 0)
            foy_no = a['foy_no']
            musteri_map = {r['foy_no']: r['musteri'] for r in db.get_rezervasyonlar()}
            musteri = musteri_map.get(foy_no, '')

            # Yevmiyedeki Adisyon Geliri kaydını güncelle
            conn_muh = mdb.get_conn()
            conn_muh.execute(
                "UPDATE yevmiye SET tutar=? WHERE islem_tipi='Adisyon Geliri' AND aciklama LIKE ?",
                (yeni_tutar, f'Föy#{foy_no} Adis#{adisyon_no}%')
            )
            conn_muh.execute(
                "UPDATE yevmiye SET tutar=? WHERE islem_tipi='Adisyon Geliri' AND aciklama=?",
                (yeni_tutar, f'Föy#{foy_no} {musteri} adisyon')
            )
            conn_muh.commit(); conn_muh.close()

        # Adisyonu güncelle
        db.update_adisyon(adisyon_no, yeni_tutar, odeme)

        # Rezervasyon adisyon toplamını yeniden hesapla (fark yerine toplam)
        if a:
            foy_no = a['foy_no']
            conn_db2 = db.get_conn()
            toplam = conn_db2.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM adisyonlar WHERE foy_no=?", (foy_no,)
            ).fetchone()[0]
            tah = conn_db2.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM adisyon_odemeler WHERE foy_no=?", (foy_no,)
            ).fetchone()[0]
            conn_db2.execute(
                "UPDATE rezervasyonlar SET adisyon=?, adis_bakiye=? WHERE foy_no=?",
                (toplam, toplam - tah, foy_no)
            )
            conn_db2.commit(); conn_db2.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/adisyon/sil', methods=['POST'])
def api_adisyon_sil():
    try:
        adisyon_no = int(request.get_json()['adisyon_no'])
        # Adisyon bilgisini al
        conn_db = db.get_conn()
        a = conn_db.execute("SELECT * FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
        conn_db.close()
        if a:
            a = dict(a)
            foy_no = a['foy_no']
            tutar = float(a['tutar'] or 0)
            otel = a.get('otel', 'GENEL')
            musteri_map = {r['foy_no']: r['musteri'] for r in db.get_rezervasyonlar()}
            musteri = musteri_map.get(foy_no, '')
            # Yevmiyeden Adisyon Geliri ve Tahsilat kayıtlarını sil
            conn_muh = mdb.get_conn()
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi='Adisyon Geliri'",
                (f'Föy#{foy_no}%Adis#{adisyon_no}%',)
            )
            # Adis no olmayan eski kayıtlar için de dene
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama=? AND islem_tipi='Adisyon Geliri'",
                (f'Föy#{foy_no} {musteri} adisyon',)
            )
            # Adisyon tahsilat kayıtlarını da sil
            conn_muh.execute(
                "DELETE FROM yevmiye WHERE aciklama LIKE ? AND islem_tipi LIKE 'Adisyon Tahsilat%'",
                (f'Föy#{foy_no} Adis#{adisyon_no}%',)
            )
            conn_muh.commit(); conn_muh.close()
            # adisyon_odemeler tablosundan sil
            conn_db2 = db.get_conn()
            conn_db2.execute("DELETE FROM adisyon_odemeler WHERE adisyon_no=?", (adisyon_no,))
            conn_db2.commit(); conn_db2.close()
            # Rezervasyon adisyon toplamını güncelle
            rez_list = db.get_rezervasyonlar()
            r = next((x for x in rez_list if x['foy_no'] == foy_no), None)
            if r:
                odenen = float(a['odenen_tutar'] or 0)
                conn_db3 = db.get_conn()
                conn_db3.execute(
                    "UPDATE rezervasyonlar SET adisyon=adisyon-?, adis_tahsilat=adis_tahsilat-?, adis_bakiye=adis_bakiye-? WHERE foy_no=?",
                    (tutar, odenen, tutar - odenen, foy_no)
                )
                conn_db3.commit(); conn_db3.close()
        # Adisyonu sil
        db.delete_adisyon(adisyon_no)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/yedek/db')
@login_required
def yedek_db():
    """Hem otel hem muhasebe veritabanlarını ZIP olarak indir."""
    import shutil, tempfile, zipfile, os
    from datetime import date
    dosya_adi = f"yedek_{date.today().isoformat()}.zip"
    tmp_zip = tempfile.NamedTemporaryFile(suffix='.zip', delete=False)
    tmp_zip.close()
    with zipfile.ZipFile(tmp_zip.name, 'w', zipfile.ZIP_DEFLATED) as zf:
        otel_db = str(db.DB_PATH)
        muh_db  = str(mdb.DB_PATH)
        if os.path.exists(otel_db):
            zf.write(otel_db, 'otel.db')
        if os.path.exists(muh_db):
            zf.write(muh_db, 'muhasebe.db')
    return send_file(tmp_zip.name, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/zip')

@app.route('/yedek/excel')
@login_required
def yedek_excel():
    """Tüm otel ve muhasebe verilerini Excel olarak indir — Power Query uyumlu."""
    import tempfile
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    HEADER_FONT  = Font(bold=True, color='FFFFFF')
    HEADER_FILL  = PatternFill('solid', fgColor='0d1b2a')
    HEADER2_FILL = PatternFill('solid', fgColor='1a2f4a')

    def make_header(ws, headers, fill=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = HEADER_FONT
            c.fill = fill or HEADER_FILL
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 35)

    # ── 1. Rezervasyon Girişleri (Otel_Oda_Durumu ile uyumlu) ────────────────
    ws1 = wb.active
    ws1.title = 'Rezervasyon Girişleri'
    h1 = ['Oda No','Otel','Föy No','Acente','Müşteri Adı','Yetişkin','Çocuk',
          'Ek Yatak','Gün Fiyat','Giriş Tarihi','Çıkış Tarihi','Toplam Gün',
          'Toplam Fiyat','Kapora','Kapora Tarihi','Tahsilat','Ödeme Türü',
          'REZ. Bakiye','Adisyon','Adisyon Tahsilat','Ödeme Türü2','ADS Bakiye','Açıklama']
    make_header(ws1, h1)
    for i, r in enumerate(db.get_rezervasyonlar(), 2):
        vals = [r.get('oda_no'), r.get('otel'), r.get('foy_no'), r.get('kanal'),
                r.get('musteri'), r.get('yetiskin'), r.get('cocuk'), r.get('ek_yatak'),
                r.get('gun_fiyat'), r.get('giris'), r.get('cikis'), r.get('toplam_gun'),
                r.get('toplam_fiyat'), r.get('kapora'), r.get('kapora_tarihi'),
                r.get('rez_tahsilat'), r.get('rez_odeme_sekli'), r.get('rez_bakiye'),
                r.get('adisyon'), r.get('adis_tahsilat'), r.get('adis_odeme_sekli'),
                r.get('adis_bakiye'), r.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws1.cell(row=i, column=col, value=v)

    # ── 2. Adisyonlar ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Adisyonlar')
    h2 = ['Adisyon No','Föy No','Oda No','Tarih','Tutar','Ödeme','Otel','Açıklama']
    make_header(ws2, h2)
    for i, a in enumerate(db.get_adisyonlar(), 2):
        vals = [a.get('adisyon_no'), a.get('foy_no'), a.get('oda_no'),
                a.get('tarih'), a.get('tutar'), a.get('odeme'),
                a.get('otel'), a.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=v)

    # ── 3. YEVMİYE (Muhasebe_Sablonu ile uyumlu) ─────────────────────────────
    ws3 = wb.create_sheet('YEVMİYE')
    h3 = ['BELGE NO','TARİH','İŞLEM TİPİ','BORÇ HESABI','ALACAK HESABI',
          'TUTAR (TL)','AÇIKLAMA','OTEL','FATURA NO']
    make_header(ws3, h3, HEADER2_FILL)
    conn = mdb.get_conn()
    yevmiye = [dict(r) for r in conn.execute(
        "SELECT * FROM yevmiye ORDER BY tarih, id").fetchall()]
    for i, r in enumerate(yevmiye, 2):
        # Hesap adını da ekle
        borc_ad = conn.execute("SELECT ad FROM hesaplar WHERE kod=?", (r.get('borc_hesap',''),)).fetchone()
        alacak_ad = conn.execute("SELECT ad FROM hesaplar WHERE kod=?", (r.get('alacak_hesap',''),)).fetchone()
        vals = [r.get('belge_no'), r.get('tarih'), r.get('islem_tipi'),
                f"{r.get('borc_hesap','')} — {borc_ad['ad'] if borc_ad else ''}",
                f"{r.get('alacak_hesap','')} — {alacak_ad['ad'] if alacak_ad else ''}",
                r.get('tutar'), r.get('aciklama'), r.get('otel'), r.get('fatura_no')]
        for col, v in enumerate(vals, 1):
            ws3.cell(row=i, column=col, value=v)

    # ── 4. PERSONEL ───────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('PERSONEL')
    h4 = ['Ad Soyad','İşe Giriş','Görev','Net Maaş','Banka IBAN',
          'TARİH','DÖNEM YIL','DÖNEM AY','NET ÖDEME','ÖDEME BANKASI','AÇIKLAMA','OTEL']
    make_header(ws4, h4, HEADER2_FILL)
    personel = mdb.get_personel()
    maaslar = [dict(r) for r in conn.execute("""
        SELECT pm.*, p.ad_soyad FROM personel_maas pm
        JOIN personel p ON pm.personel_id=p.id ORDER BY pm.tarih
    """).fetchall()]
    max_rows = max(len(personel), len(maaslar), 1)
    for i in range(max_rows):
        row_idx = i + 2
        if i < len(personel):
            p = personel[i]
            ws4.cell(row=row_idx, column=1, value=p.get('ad_soyad'))
            ws4.cell(row=row_idx, column=2, value=p.get('ise_giris'))
            ws4.cell(row=row_idx, column=3, value=p.get('gorev'))
            ws4.cell(row=row_idx, column=4, value=p.get('net_maas'))
            ws4.cell(row=row_idx, column=5, value=p.get('banka_iban'))
        if i < len(maaslar):
            m = maaslar[i]
            ws4.cell(row=row_idx, column=6,  value=m.get('tarih'))
            ws4.cell(row=row_idx, column=7,  value=m.get('donem_yil'))
            ws4.cell(row=row_idx, column=8,  value=m.get('donem_ay'))
            ws4.cell(row=row_idx, column=9,  value=m.get('net_odeme'))
            ws4.cell(row=row_idx, column=10, value=m.get('odeme_banka'))
            ws4.cell(row=row_idx, column=11, value=m.get('aciklama'))
            ws4.cell(row=row_idx, column=12, value=m.get('otel'))

    # ── 5. STOK ───────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet('STOK')
    h5 = ['TARİH','BELGE NO','AÇIKLAMA','KATEGORİ','TUTAR','ÖDEME HESABI','FATURA','OTEL','NOT']
    make_header(ws5, h5, HEADER2_FILL)
    stok = [dict(r) for r in conn.execute("SELECT * FROM stok ORDER BY tarih").fetchall()]
    for i, r in enumerate(stok, 2):
        vals = [r.get('tarih'), r.get('belge_no'), r.get('aciklama'), r.get('kategori'),
                r.get('tutar'), r.get('odeme_hesap'), r.get('fatura_var'), r.get('otel'), r.get('not_')]
        for col, v in enumerate(vals, 1):
            ws5.cell(row=i, column=col, value=v)

    # ── 6. VERGİ ─────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet('VERGİ')
    h6 = ['TARİH','DÖNEM YIL','DÖNEM AY','VERGİ TÜRÜ','MATRAH','TUTAR','ÖDEME BANKASI','DURUM','AÇIKLAMA']
    make_header(ws6, h6, HEADER2_FILL)
    vergi = [dict(r) for r in conn.execute("SELECT * FROM vergi ORDER BY donem_yil, donem_ay").fetchall()]
    for i, r in enumerate(vergi, 2):
        vals = [r.get('tarih'), r.get('donem_yil'), r.get('donem_ay'), r.get('vergi_turu'),
                r.get('matrah'), r.get('tutar'), r.get('odeme_banka'), r.get('durum'), r.get('aciklama')]
        for col, v in enumerate(vals, 1):
            ws6.cell(row=i, column=col, value=v)

    # ── 7. ACENTE CARİ ────────────────────────────────────────────────────────
    ws7 = wb.create_sheet('ACENTE CARİ')
    h7 = ['TARİH','ACENTE','FÖY NO','REZ NO','MİSAFİR','REZ TUTARI',
          'KOM ORAN','KOMİSYON TL','GELEN ÖDEME','OTEL']
    make_header(ws7, h7, HEADER2_FILL)
    acente = [dict(r) for r in conn.execute("SELECT * FROM acente_cari ORDER BY tarih").fetchall()]
    for i, r in enumerate(acente, 2):
        vals = [r.get('tarih'), r.get('acente_kod'), r.get('foy_no'), r.get('rez_no'),
                r.get('misafir'), r.get('rez_tutari'), r.get('komisyon_oran'),
                r.get('komisyon_tl'), r.get('gelen_odeme'), r.get('otel')]
        for col, v in enumerate(vals, 1):
            ws7.cell(row=i, column=col, value=v)

    # ── 8. ORTAK CARİ ─────────────────────────────────────────────────────────
    ws8 = wb.create_sheet('ORTAK CARİ')
    h8 = ['TARİH','ORTAK','BELGE NO','AÇIKLAMA','KATEGORİ','TUTAR','ÖDEME','İADE','NET','OTEL']
    make_header(ws8, h8, HEADER2_FILL)
    ortak = [dict(r) for r in conn.execute("SELECT * FROM ortak_cari ORDER BY tarih").fetchall()]
    for i, r in enumerate(ortak, 2):
        net = (r.get('tutar') or 0) - (r.get('iade') or 0)
        vals = [r.get('tarih'), r.get('ortak'), r.get('belge_no'), r.get('aciklama'),
                r.get('gider_kategori'), r.get('tutar'), r.get('odeme_sekli'),
                r.get('iade'), net, r.get('otel')]
        for col, v in enumerate(vals, 1):
            ws8.cell(row=i, column=col, value=v)

    # ── 9. MİZAN ─────────────────────────────────────────────────────────────
    ws9 = wb.create_sheet('MİZAN')
    h9 = ['KOD','HESAP ADI','TİP','GRUP','BORÇ','ALACAK','BAKİYE']
    make_header(ws9, h9, HEADER2_FILL)
    hesaplar = [dict(r) for r in conn.execute("SELECT * FROM hesaplar WHERE aktif=1 ORDER BY kod").fetchall()]
    yil = date.today().year
    row_idx = 2
    for h in hesaplar:
        borc = conn.execute("SELECT COALESCE(SUM(tutar),0) FROM yevmiye WHERE yil=? AND borc_hesap=?",
                           (yil, h['kod'])).fetchone()[0] or 0
        alacak = conn.execute("SELECT COALESCE(SUM(tutar),0) FROM yevmiye WHERE yil=? AND alacak_hesap=?",
                             (yil, h['kod'])).fetchone()[0] or 0
        if borc == 0 and alacak == 0:
            continue
        bakiye = alacak - borc if h['tip'] in ('Gelir','Pasif','Ozkaynak') else borc - alacak
        vals = [h['kod'], h['ad'], h['tip'], h['grup'], borc, alacak, bakiye]
        for col, v in enumerate(vals, 1):
            ws9.cell(row=row_idx, column=col, value=v)
        row_idx += 1

    conn.close()

    # Sütun genişliklerini ayarla
    for ws in [ws1,ws2,ws3,ws4,ws5,ws6,ws7,ws8,ws9]:
        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 40)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    wb.save(tmp.name)

    dosya_adi = f"otel_yedek_{date.today().isoformat()}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=dosya_adi,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/import', methods=['POST'])
def api_import():
    """Excel dosyasını upload edip SQLite'a aktar."""
    try:
        f = request.files.get('excel')
        if not f:
            return jsonify({'ok': False, 'error': 'Dosya seçilmedi'}), 400
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        rez_count, adis_count = db.import_from_excel(tmp_path)
        os.unlink(tmp_path)
        return jsonify({'ok': True, 'rezervasyon': rez_count, 'adisyon': adis_count})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"🌐 Sunucu → http://localhost:{port}")
    app.run(debug=False, host='0.0.0.0', port=port)
