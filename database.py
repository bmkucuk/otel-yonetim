#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SQLite veritabanı katmanı — Otel Leo & Cunda Villa Web"""
import sqlite3
import os
from datetime import date, datetime

_data_dir = "/data" if os.path.isdir("/data") else "."
DB_PATH = os.environ.get("DB_PATH", os.path.join(_data_dir, "otel.db"))


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_conn()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS rezervasyonlar (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        oda_no          INTEGER NOT NULL,
        otel            TEXT NOT NULL DEFAULT 'LEO',
        foy_no          INTEGER UNIQUE NOT NULL,
        kanal           TEXT DEFAULT '',
        musteri         TEXT NOT NULL,
        yetiskin        INTEGER DEFAULT 1,
        cocuk           INTEGER DEFAULT 0,
        ek_yatak        TEXT DEFAULT 'Yok',
        gun_fiyat       REAL DEFAULT 0,
        giris           TEXT,
        cikis           TEXT,
        toplam_gun      INTEGER DEFAULT 0,
        toplam_fiyat    REAL DEFAULT 0,
        kapora          REAL DEFAULT 0,
        kapora_tarihi   TEXT,
        rez_tahsilat    REAL DEFAULT 0,
        rez_odeme_sekli TEXT DEFAULT '',
        rez_bakiye      REAL DEFAULT 0,
        adisyon         REAL DEFAULT 0,
        adis_tahsilat   REAL DEFAULT 0,
        adis_odeme_sekli TEXT DEFAULT '',
        adis_bakiye     REAL DEFAULT 0,
        aciklama        TEXT DEFAULT '',
        created_at      TEXT DEFAULT (datetime('now','localtime')),
        updated_at      TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS adisyonlar (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        adisyon_no  INTEGER UNIQUE NOT NULL,
        foy_no      INTEGER NOT NULL,
        oda_no      INTEGER,
        tarih       TEXT,
        tutar       REAL DEFAULT 0,
        odeme       TEXT DEFAULT 'Oda Hesabına',
        aciklama    TEXT DEFAULT '',
        otel        TEXT DEFAULT '',
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS adisyon_odemeler (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        adisyon_no  INTEGER NOT NULL,
        foy_no      INTEGER NOT NULL,
        tarih       TEXT,
        tutar       REAL DEFAULT 0,
        odeme_sekli TEXT DEFAULT '',
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS odalar (
        oda_no  INTEGER PRIMARY KEY,
        otel    TEXT NOT NULL,
        tip     TEXT DEFAULT 'Standart'
    );
    """)
    # Oda listesini doldur
    count = conn.execute("SELECT COUNT(*) FROM odalar").fetchone()[0]
    if count == 0:
        leo = [(i, 'LEO') for i in range(11, 30)]
        cv  = [(i, 'CV')  for i in range(1, 11)]
        conn.executemany("INSERT OR IGNORE INTO odalar(oda_no,otel) VALUES(?,?)", leo + cv)
    # Migration: durum kolonu yoksa ekle
    cols = [r[1] for r in conn.execute("PRAGMA table_info(rezervasyonlar)").fetchall()]
    if 'checkin' not in cols:
        conn.execute("ALTER TABLE rezervasyonlar ADD COLUMN checkin INTEGER DEFAULT 0")
        conn.execute("UPDATE rezervasyonlar SET checkin=0")
    if 'durum' not in cols:
        conn.execute("ALTER TABLE rezervasyonlar ADD COLUMN durum TEXT DEFAULT 'Aktif'")
        conn.execute("UPDATE rezervasyonlar SET durum='Aktif' WHERE durum IS NULL")
    # Migration: adisyonlar tablosuna odeme kolonları ekle
    adis_cols = [r[1] for r in conn.execute("PRAGMA table_info(adisyonlar)").fetchall()]
    if 'odendi' not in adis_cols:
        conn.execute("ALTER TABLE adisyonlar ADD COLUMN odendi INTEGER DEFAULT 0")
    if 'odeme_tarihi' not in adis_cols:
        conn.execute("ALTER TABLE adisyonlar ADD COLUMN odeme_tarihi TEXT DEFAULT ''")
    if 'odeme_sekli' not in adis_cols:
        conn.execute("ALTER TABLE adisyonlar ADD COLUMN odeme_sekli TEXT DEFAULT ''")
    if 'odenen_tutar' not in adis_cols:
        conn.execute("ALTER TABLE adisyonlar ADD COLUMN odenen_tutar REAL DEFAULT 0")
        # Mevcut tam ödenmişlerin odenen_tutar'ını tutar'a eşitle
        conn.execute("UPDATE adisyonlar SET odenen_tutar=tutar WHERE odendi=1")
    conn.commit()
    conn.close()


# ── Rezervasyon CRUD ──────────────────────────────────────────────────────────

def get_rezervasyonlar(q=None, otel=None):
    conn = get_conn()
    sql = "SELECT * FROM rezervasyonlar WHERE 1=1"
    params = []
    if otel and otel != 'Tümü':
        sql += " AND otel=?"; params.append(otel)
    if q:
        sql += " AND (LOWER(musteri) LIKE ? OR CAST(foy_no AS TEXT) LIKE ? OR CAST(oda_no AS TEXT) LIKE ?)"
        params += [f'%{q.lower()}%', f'%{q}%', f'%{q}%']
    sql += " ORDER BY foy_no"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_rezervasyon(foy_no):
    conn = get_conn()
    r = conn.execute("SELECT * FROM rezervasyonlar WHERE foy_no=?", (foy_no,)).fetchone()
    conn.close()
    return dict(r) if r else None

def save_rezervasyon(data):
    conn = get_conn()
    # Toplam gün ve fiyat hesapla
    giris = data.get('giris')
    cikis = data.get('cikis')
    gun_fiyat = float(data.get('gun_fiyat') or 0)
    if giris and cikis:
        from datetime import date as d_
        g = d_.fromisoformat(giris)
        c = d_.fromisoformat(cikis)
        toplam_gun = (c - g).days
        toplam_fiyat = toplam_gun * gun_fiyat
    else:
        toplam_gun = 0
        toplam_fiyat = 0
    kapora = float(data.get('kapora') or 0)
    rez_tahsilat = float(data.get('rez_tahsilat') or 0)
    rez_bakiye = max(0, toplam_fiyat - kapora - rez_tahsilat)

    conn.execute("""
        INSERT INTO rezervasyonlar
            (oda_no,otel,foy_no,kanal,musteri,yetiskin,cocuk,ek_yatak,
             gun_fiyat,giris,cikis,toplam_gun,toplam_fiyat,
             kapora,kapora_tarihi,rez_tahsilat,rez_odeme_sekli,rez_bakiye,aciklama)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        int(data['oda_no']), data['otel'], int(data['foy_no']),
        data.get('kanal',''), data['musteri'],
        int(data.get('yetiskin',1)), int(data.get('cocuk',0)),
        data.get('ek_yatak','Yok'), gun_fiyat,
        giris, cikis, toplam_gun, toplam_fiyat,
        kapora, data.get('kapora_tarihi'),
        rez_tahsilat, data.get('rez_odeme_sekli',''), rez_bakiye,
        data.get('aciklama','')
    ))
    conn.commit(); conn.close()

def update_rezervasyon(foy_no, data):
    conn = get_conn()
    giris = data.get('giris')
    cikis = data.get('cikis')
    gun_fiyat = float(data.get('gun_fiyat') or 0)
    if giris and cikis:
        from datetime import date as d_
        g = d_.fromisoformat(giris)
        c = d_.fromisoformat(cikis)
        toplam_gun = (c - g).days
        toplam_fiyat = toplam_gun * gun_fiyat
    else:
        toplam_gun = int(data.get('toplam_gun') or 0)
        toplam_fiyat = float(data.get('toplam_fiyat') or 0)
    kapora = float(data.get('kapora') or 0)
    conn.execute("""
        UPDATE rezervasyonlar SET
            oda_no=?, otel=?, kanal=?, musteri=?, yetiskin=?, cocuk=?,
            ek_yatak=?, gun_fiyat=?, giris=?, cikis=?, toplam_gun=?,
            toplam_fiyat=?, kapora=?, kapora_tarihi=?, aciklama=?,
            updated_at=datetime('now','localtime')
        WHERE foy_no=?
    """, (
        int(data['oda_no']), data['otel'], data.get('kanal',''),
        data['musteri'], int(data.get('yetiskin',1)), int(data.get('cocuk',0)),
        data.get('ek_yatak','Yok'), gun_fiyat, giris, cikis,
        toplam_gun, toplam_fiyat, kapora, data.get('kapora_tarihi'),
        data.get('aciklama',''), int(foy_no)
    ))
    conn.commit(); conn.close()

def save_rez_tahsilat(foy_no, tutar, odeme):
    conn = get_conn()
    r = conn.execute("SELECT * FROM rezervasyonlar WHERE foy_no=?", (foy_no,)).fetchone()
    if not r: conn.close(); raise RuntimeError(f"Föy #{foy_no} bulunamadı")
    yeni = (r['rez_tahsilat'] or 0) + tutar
    bakiye = max(0, (r['toplam_fiyat'] or 0) - (r['kapora'] or 0) - yeni)
    conn.execute("""
        UPDATE rezervasyonlar SET rez_tahsilat=?, rez_odeme_sekli=?, rez_bakiye=?,
        updated_at=datetime('now','localtime') WHERE foy_no=?
    """, (yeni, odeme, bakiye, foy_no))
    conn.commit(); conn.close()

def save_adis_tahsilat(foy_no, tutar, odeme):
    conn = get_conn()
    r = conn.execute("SELECT * FROM rezervasyonlar WHERE foy_no=?", (foy_no,)).fetchone()
    if not r: conn.close(); raise RuntimeError(f"Föy #{foy_no} bulunamadı")
    yeni = (r['adis_tahsilat'] or 0) + tutar
    bakiye = max(0, (r['adisyon'] or 0) - yeni)
    conn.execute("""
        UPDATE rezervasyonlar SET adis_tahsilat=?, adis_odeme_sekli=?, adis_bakiye=?,
        updated_at=datetime('now','localtime') WHERE foy_no=?
    """, (yeni, odeme, bakiye, foy_no))
    conn.commit(); conn.close()

def save_adisyon_odeme(adisyon_no, foy_no, tutar, odeme_sekli):
    """Adisyon bazlı kısmi ödeme kaydeder."""
    tarih = __import__('datetime').date.today().isoformat()
    conn = get_conn()
    # Ödeme kaydı ekle
    conn.execute(
        "INSERT INTO adisyon_odemeler (adisyon_no, foy_no, tarih, tutar, odeme_sekli) VALUES (?,?,?,?,?)",
        (adisyon_no, foy_no, tarih, tutar, odeme_sekli)
    )
    # Adisyonun odenen_tutar'ını güncelle
    a = conn.execute("SELECT tutar, odenen_tutar FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
    if a:
        yeni_odenen = (a['odenen_tutar'] or 0) + tutar
        tam_odendi = yeni_odenen >= a['tutar']
        conn.execute(
            "UPDATE adisyonlar SET odenen_tutar=?, odendi=?, odeme_tarihi=?, odeme_sekli=? WHERE adisyon_no=?",
            (yeni_odenen, 1 if tam_odendi else 0, tarih if tam_odendi else '', odeme_sekli if tam_odendi else '', adisyon_no)
        )
    conn.commit(); conn.close()
    return tarih

def get_adisyon_odemeler(adisyon_no):
    """Bir adisyona ait tüm ödemeleri döndürür."""
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM adisyon_odemeler WHERE adisyon_no=? ORDER BY id",
        (adisyon_no,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_next_foy_no():
    conn = get_conn()
    row = conn.execute("SELECT MAX(foy_no) as m FROM rezervasyonlar").fetchone()
    conn.close()
    return (row['m'] or 0) + 1


# ── Adisyon CRUD ──────────────────────────────────────────────────────────────

def get_adisyonlar(foy_no=None):
    conn = get_conn()
    if foy_no:
        rows = conn.execute("SELECT * FROM adisyonlar WHERE foy_no=? ORDER BY adisyon_no", (foy_no,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM adisyonlar ORDER BY adisyon_no DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_next_adisyon_no():
    conn = get_conn()
    row = conn.execute("SELECT MAX(adisyon_no) as m FROM adisyonlar").fetchone()
    conn.close()
    return (row['m'] or 0) + 1

def save_adisyon(data):
    conn = get_conn()
    foy_no = int(data['foy_no'])
    rez = conn.execute("SELECT * FROM rezervasyonlar WHERE foy_no=?", (foy_no,)).fetchone()
    conn.execute("""
        INSERT INTO adisyonlar (adisyon_no,foy_no,oda_no,tarih,tutar,odeme,aciklama,otel)
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        int(data['adisyon_no']), foy_no,
        rez['oda_no'] if rez else None,
        data.get('tarih'), float(data['tutar']),
        data.get('odeme','Oda Hesabına'), data.get('aciklama',''),
        rez['otel'] if rez else ''
    ))
    # Rezervasyondaki adisyon toplamını güncelle
    _sync_adisyon_totals(conn, foy_no)
    conn.commit(); conn.close()

def update_adisyon(adisyon_no, tutar, odeme):
    conn = get_conn()
    a = conn.execute("SELECT foy_no FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
    conn.execute("UPDATE adisyonlar SET tutar=?, odeme=? WHERE adisyon_no=?",
                 (tutar, odeme, adisyon_no))
    if a:
        _sync_adisyon_totals(conn, a['foy_no'])
    conn.commit(); conn.close()

def delete_rezervasyon(foy_no):
    conn = get_conn()
    conn.execute("DELETE FROM rezervasyonlar WHERE foy_no=?", (foy_no,))
    conn.commit(); conn.close()

def delete_adisyon(adisyon_no):
    conn = get_conn()
    a = conn.execute("SELECT foy_no FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,)).fetchone()
    conn.execute("DELETE FROM adisyonlar WHERE adisyon_no=?", (adisyon_no,))
    if a:
        _sync_adisyon_totals(conn, a['foy_no'])
    conn.commit(); conn.close()

def _sync_adisyon_totals(conn, foy_no):
    """Adisyon ekle/sil/güncelle sonrası rezervasyondaki adisyon özetini hesapla."""
    NAKIT = ('Nakit', 'Kredi Kartı', 'Havale', 'EFT')
    rows = conn.execute("SELECT tutar, odeme FROM adisyonlar WHERE foy_no=?", (foy_no,)).fetchall()
    toplam  = sum(r['tutar'] or 0 for r in rows)
    tahsil  = sum(r['tutar'] or 0 for r in rows if r['odeme'] in NAKIT)
    bakiye  = max(0, toplam - tahsil)
    conn.execute("""
        UPDATE rezervasyonlar SET adisyon=?, adis_tahsilat=?, adis_bakiye=?,
        updated_at=datetime('now','localtime') WHERE foy_no=?
    """, (toplam, tahsil, bakiye, foy_no))


# ── Dashboard queries ─────────────────────────────────────────────────────────

def get_dashboard(today_str):
    conn = get_conn()
    girisler = [dict(r) for r in conn.execute(
        "SELECT * FROM rezervasyonlar WHERE giris=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY oda_no", (today_str,)).fetchall()]
    cikislar = [dict(r) for r in conn.execute(
        "SELECT * FROM rezervasyonlar WHERE cikis=? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY oda_no", (today_str,)).fetchall()]
    aktifler = [dict(r) for r in conn.execute(
        "SELECT * FROM rezervasyonlar WHERE giris<=? AND cikis>? AND (durum IS NULL OR durum != 'Kapora Yandı') ORDER BY oda_no",
        (today_str, today_str)).fetchall()]
    kahvalti = conn.execute(
        "SELECT SUM(yetiskin+cocuk) FROM rezervasyonlar WHERE giris<=? AND cikis>=? AND (durum IS NULL OR durum != 'Kapora Yandı')",
        (today_str, today_str)).fetchone()[0] or 0
    conn.close()
    return girisler, cikislar, aktifler, int(kahvalti)


# ── Excel import ──────────────────────────────────────────────────────────────

def import_from_excel(excel_path):
    """Excel'den SQLite'a toplu aktarım."""
    from openpyxl import load_workbook
    from datetime import datetime, date, timedelta

    def _to_date(val):
        if val is None: return None
        if isinstance(val, datetime): return val.date().isoformat()
        if isinstance(val, date): return val.isoformat()
        if isinstance(val, int):
            try: return (datetime(1899,12,30)+timedelta(days=val)).date().isoformat()
            except: return None
        return None

    def _f(val):
        if val is None: return 0.0
        if isinstance(val, (int,float)): return float(val)
        return 0.0

    wb = load_workbook(excel_path, data_only=True)
    ws  = wb['Rezervasyon Girişleri']
    ws2 = wb['Adisyonlar']

    conn = get_conn()
    rez_count = 0
    adis_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        def col(i): return row[i] if len(row)>i else None

        giris = _to_date(col(9))
        cikis = _to_date(col(10))
        if giris and giris[:4] < '2000': giris = None
        if cikis and cikis[:4] < '2000': cikis = None
        gun_fiyat = _f(col(8))
        if giris and cikis:
            from datetime import date as d_
            g = d_.fromisoformat(giris); c = d_.fromisoformat(cikis)
            toplam_gun = (c-g).days
            toplam_fiyat = toplam_gun * gun_fiyat
        else:
            toplam_gun = int(_f(col(11)))
            toplam_fiyat = _f(col(12))

        kapora       = _f(col(13))
        rez_tah      = _f(col(15))
        _rez_bak_raw = col(17)
        rez_bak = float(_rez_bak_raw) if isinstance(_rez_bak_raw,(int,float)) else max(0,toplam_fiyat-kapora-rez_tah)
        adisyon      = _f(col(18))
        adis_tah     = _f(col(19))
        _adis_bak    = col(21)
        adis_bak = float(_adis_bak) if isinstance(_adis_bak,(int,float)) else max(0,adisyon-adis_tah)

        try:
            conn.execute("""
                INSERT OR REPLACE INTO rezervasyonlar
                (oda_no,otel,foy_no,kanal,musteri,yetiskin,cocuk,ek_yatak,
                 gun_fiyat,giris,cikis,toplam_gun,toplam_fiyat,
                 kapora,kapora_tarihi,rez_tahsilat,rez_odeme_sekli,rez_bakiye,
                 adisyon,adis_tahsilat,adis_odeme_sekli,adis_bakiye,aciklama)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                col(0), str(col(1) or ('LEO' if (col(0) or 0)>10 else 'CV')),
                col(2), str(col(3) or ''), str(col(4) or ''),
                int(_f(col(5)) or 1), int(_f(col(6)) or 0),
                str(col(7) or 'Yok'), gun_fiyat,
                giris, cikis, toplam_gun, toplam_fiyat,
                kapora, _to_date(col(14)),
                rez_tah, str(col(16) or ''), rez_bak,
                adisyon, adis_tah, str(col(20) or ''), adis_bak,
                str(col(22) or '')
            ))
            rez_count += 1
        except Exception as e:
            print(f"  Rezervasyon atlandı (foy={col(2)}): {e}")

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        try:
            conn.execute("""
                INSERT OR REPLACE INTO adisyonlar
                (adisyon_no,foy_no,oda_no,tarih,tutar,odeme,otel)
                VALUES(?,?,?,?,?,?,?)
            """, (row[0], row[1], row[2], _to_date(row[3]),
                  _f(row[4]), str(row[5] or ''), str(row[6] or '')))
            adis_count += 1
        except Exception as e:
            print(f"  Adisyon atlandı: {e}")

    conn.commit(); conn.close(); wb.close()
    return rez_count, adis_count
